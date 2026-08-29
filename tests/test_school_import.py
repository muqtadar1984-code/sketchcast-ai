"""The onboarding importer's validation, exercised against the real templates.

Only parse.py is covered here, and deliberately so: it is the half that decides
whether a school's file is safe to write, and it is pure, so it runs locally
where `worker/` and `supabase` cannot even be imported. apply.py is the half
that touches the network and is reviewed rather than unit-tested — the same
split the rest of this repo uses.

The fixtures below FILL the actual shipped templates rather than hand-rolling a
workbook, so a column renamed in the template breaks these tests. That is the
point: the headings are the contract between a school's spreadsheet and the
database.
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from tools.school_import.parse import parse_people, parse_setup  # noqa: E402

# The templates are versioned WITH the parser: their column headings are the
# contract between a school's spreadsheet and the database, so a heading that
# changes must break these tests.
TEMPLATES = Path(__file__).resolve().parents[1] / "tools" / "templates"
SETUP_T = TEMPLATES / "SketchCast_School_Setup_TEMPLATE.xlsx"
PEOPLE_T = TEMPLATES / "SketchCast_Students_Parents_TEMPLATE.xlsx"

TEACHER = "aisha.rahman@greenfield.edu.my"
HEAD = "priya.menon@greenfield.edu.my"


def _fill(src: Path, dest: Path, sheets: dict[str, list[list]], form: dict | None = None):
    """Write rows under each tab's headings, clearing the amber example row."""
    from openpyxl import load_workbook

    wb = load_workbook(src)
    if form:
        ws = wb["1. School"]
        for row in ws.iter_rows(min_col=1, max_col=2):
            label = str(row[0].value or "").strip().rstrip("*").strip()
            if label in form:
                row[1].value = form[label]
    for title, rows in sheets.items():
        ws = wb[title]
        ws.delete_rows(2, ws.max_row)  # drop the example + blank tinted rows
        for r in rows:
            ws.append(r)
    wb.save(dest)
    return dest


def _setup(tmp_path, **over):
    sheets = {
        "2. Periods": over.pop("periods", [[1, "P1", "08:45"], [2, "P2", "09:30"], [3, "P3", "10:15"]]),
        "3. Breaks": over.pop("breaks", [[2, "Recess", "10:00", 15]]),
        "4. Staff": over.pop("staff", [
            ["Priya Menon", HEAD, "school_admin", "", "en", "T-1", ""],
            ["Aisha Rahman", TEACHER, "teacher", "Science", "en", "T-2", ""],
        ]),
        "5. Classes": over.pop("classes", [["Grade 7A", "Grade 7", TEACHER, "B-204", ""]]),
        "6. Coordinator Scope": over.pop("scopes", []),
        "7. Timetable": over.pop("slots", [["Grade 7A", "1 Monday", 1, "Science", TEACHER, "B-204"]]),
    }
    form = {
        "School name": "Greenfield International School",
        "Preferred web address": "greenfield",
        "Country": "MY", "Default interface language": "en",
        "Teaching days per week": 5, "School day starts": "08:45",
        "School day ends": "15:45", "Length of one period, minutes": 45,
    }
    form.update(over.pop("form", {}))
    return _fill(SETUP_T, tmp_path / "setup.xlsx", sheets, form)


def _people(tmp_path, **over):
    sheets = {
        "1. Students": over.pop("students", [
            ["Arif Hassan", "arif.hassan", "Grade 7A", "2026-0431", "", "en", ""],
        ]),
        "2. Parents": over.pop("parents", [
            ["Nadia Hassan", "n.hassan@example.com", "Mother", "en", "", ""],
        ]),
        "3. Parent Links": over.pop("links", [["n.hassan@example.com", "arif.hassan", ""]]),
    }
    return _fill(PEOPLE_T, tmp_path / "people.xlsx", sheets)


class TestAGoodFile:
    def test_it_parses_and_is_importable(self, tmp_path):
        plan = parse_setup(_setup(tmp_path))
        assert plan.ok, plan.errors
        assert plan.school["slug"] == "greenfield"
        assert plan.school["name"] == "Greenfield International School"
        assert [p["number"] for p in plan.periods] == [1, 2, 3]
        assert len(plan.staff) == 2 and len(plan.classes) == 1 and len(plan.slots) == 1

    def test_people_join_to_the_setup_file(self, tmp_path):
        plan = parse_setup(_setup(tmp_path))
        parse_people(_people(tmp_path), plan)
        assert plan.ok, plan.errors
        assert plan.students[0]["username"] == "arif.hassan"
        assert plan.links == [{"row": 2, "parent_email": "n.hassan@example.com",
                               "username": "arif.hassan"}]

    def test_the_untouched_template_imports_as_empty(self, tmp_path):
        """The example rows are amber and marked EXAMPLE. A school that forgets to
        delete one must not have a fictional teacher created."""
        plan = parse_setup(SETUP_T)
        assert plan.staff == [] and plan.classes == [] and plan.slots == []


class TestReferencesBetweenSheets:
    """The whole reason a school never sees a UUID: every cross-sheet reference is
    a human key, so every one of them can be mistyped."""

    def test_a_class_teacher_who_is_not_staff_is_refused(self, tmp_path):
        plan = parse_setup(_setup(tmp_path, classes=[["Grade 7A", "Grade 7", "ghost@x.com", "", ""]]))
        assert not plan.ok
        assert any("not on the Staff tab" in str(e) for e in plan.errors)

    def test_a_timetable_class_that_does_not_exist_is_refused(self, tmp_path):
        plan = parse_setup(_setup(tmp_path,
                                  slots=[["Grade 7 A", "1 Monday", 1, "Science", TEACHER, ""]]))
        assert not plan.ok
        # The exact trap the template warns about twice: a stray space.
        assert any("'Grade 7 A' is not on the Classes tab" in str(e) for e in plan.errors)

    def test_a_student_in_an_unknown_class_is_refused(self, tmp_path):
        plan = parse_setup(_setup(tmp_path))
        parse_people(_people(tmp_path, students=[
            ["Arif Hassan", "arif.hassan", "Grade 8B", "", "", "en", ""]]), plan)
        assert not plan.ok
        assert any("Grade 8B" in str(e) for e in plan.errors)

    def test_a_link_to_an_unknown_parent_or_child_is_refused(self, tmp_path):
        plan = parse_setup(_setup(tmp_path))
        parse_people(_people(tmp_path, links=[["nobody@example.com", "no.such.child", ""]]), plan)
        msgs = " ".join(str(e) for e in plan.errors)
        assert "not on the Parents tab" in msgs and "not on the Students tab" in msgs

    def test_a_coordinator_grant_needs_the_coordinator_role(self, tmp_path):
        # On the Staff tab as a plain teacher — the grant would silently do nothing.
        plan = parse_setup(_setup(tmp_path, scopes=[[TEACHER, "Grade 7", "Science"]]))
        assert not plan.ok
        assert any("Role is not 'coordinator'" in str(e) for e in plan.errors)


class TestUniqueness:
    def test_two_staff_cannot_share_an_email(self, tmp_path):
        plan = parse_setup(_setup(tmp_path, staff=[
            ["Priya Menon", HEAD, "school_admin", "", "en", "", ""],
            ["Someone Else", HEAD, "teacher", "", "en", "", ""]]))
        assert not plan.ok
        assert any("must be unique" in str(e) for e in plan.errors)

    def test_two_students_cannot_share_a_username(self, tmp_path):
        plan = parse_setup(_setup(tmp_path))
        parse_people(_people(tmp_path, students=[
            ["Arif Hassan", "arif.hassan", "Grade 7A", "", "", "en", ""],
            ["Arif Hussain", "arif.hassan", "Grade 7A", "", "", "en", ""]]), plan)
        assert not plan.ok
        assert any("must be unique" in str(e) for e in plan.errors)

    def test_a_class_cannot_be_in_two_places_at_once(self, tmp_path):
        plan = parse_setup(_setup(tmp_path, slots=[
            ["Grade 7A", "1 Monday", 1, "Science", TEACHER, ""],
            ["Grade 7A", "1 Monday", 1, "Maths", TEACHER, ""]]))
        assert not plan.ok
        assert any("two places at once" in str(e) for e in plan.errors)


class TestFieldShapes:
    def test_a_bad_web_address_is_refused(self, tmp_path):
        plan = parse_setup(_setup(tmp_path, form={"Preferred web address": "Green Field!"}))
        assert not plan.ok
        assert any("lowercase letters, numbers and hyphens" in str(e) for e in plan.errors)

    def test_a_bad_username_is_refused(self, tmp_path):
        plan = parse_setup(_setup(tmp_path))
        parse_people(_people(tmp_path, students=[
            ["Arif Hassan", "Arif Hassan", "Grade 7A", "", "", "en", ""]]), plan)
        assert not plan.ok
        assert any("no spaces" in str(e) for e in plan.errors)

    def test_a_staff_email_on_the_student_domain_is_refused(self, tmp_path):
        """students.sketchcast.app addresses are synthetic sign-in identities —
        a real member of staff must never be given one."""
        plan = parse_setup(_setup(tmp_path, staff=[
            ["Odd One", "someone@students.sketchcast.app", "teacher", "", "en", "", ""]]))
        assert not plan.ok
        assert any("student sign-in address" in str(e) for e in plan.errors)

    def test_times_must_be_24_hour(self, tmp_path):
        plan = parse_setup(_setup(tmp_path, periods=[[1, "P1", "8.45am"]]))
        assert not plan.ok
        assert any("24-hour HH:MM" in str(e) for e in plan.errors)

    def test_an_unsupported_locale_warns_but_does_not_block(self, tmp_path):
        plan = parse_setup(_setup(tmp_path, staff=[
            ["Priya Menon", HEAD, "school_admin", "", "de", "", ""],
            ["Aisha Rahman", TEACHER, "teacher", "Science", "en", "", ""]]))
        assert plan.ok, plan.errors
        assert any("not supported" in str(w) for w in plan.warnings)

    def test_an_overloaded_teacher_warns_but_does_not_block(self, tmp_path):
        """A school's own stated maximum, checked against its own timetable —
        worth saying out loud, not worth refusing the import over."""
        slots = [["Grade 7A", "1 Monday", n, "Science", TEACHER, ""] for n in (1, 2, 3)]
        plan = parse_setup(_setup(tmp_path, form={"Max periods per teacher per day": 2},
                                  slots=slots, periods=[[n, f"P{n}", f"{7 + n:02d}:00"] for n in (1, 2, 3)]))
        assert plan.ok, plan.errors
        assert any("above the school's stated maximum" in str(w) for w in plan.warnings)


class TestMissingStructure:
    def test_a_renamed_tab_is_a_loud_failure(self, tmp_path):
        from openpyxl import load_workbook

        p = _setup(tmp_path)
        wb = load_workbook(p)
        wb["4. Staff"].title = "Staff"
        wb.save(p)
        plan = parse_setup(p)
        assert not plan.ok
        assert any("do not rename" in str(e) for e in plan.errors)

    def test_the_timetable_tab_is_optional(self, tmp_path):
        plan = parse_setup(_setup(tmp_path, slots=[]))
        assert plan.ok, plan.errors
        assert plan.slots == []
