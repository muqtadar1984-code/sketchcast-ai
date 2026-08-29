"""Read the two onboarding workbooks and validate them into a Plan.

PURE ON PURPOSE. Nothing here imports supabase or touches the network, so the
whole of the risky part — "did the school fill this in correctly?" — is testable
locally, where `worker/` itself cannot even be imported (the documented env gap).
apply.py takes the Plan this produces and is the only thing that writes.

The workbooks are the ones in Edtech/onboarding/. Column headings ARE the
contract: they carry a trailing "*" when required, and the loader matches on the
heading text, so a renamed column is a loud failure rather than a silent skip.

Every reference between sheets is a HUMAN key — an email, a class name, a
username — because a school cannot be asked to know our UUIDs. Resolving those
to ids is apply.py's job; proving they all resolve is this file's job.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

ROLES = {"school_admin", "teacher", "coordinator"}
LOCALES = {"en", "ms", "ms-arab", "ar", "fr", "es", "pt", "hi", "mr", "te"}
# Matches src/utils/student.ts — usernames are backed by a synthetic address.
STUDENT_EMAIL_DOMAIN = "students.sketchcast.app"

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s.]+\.[^@\s]+$")
_USERNAME_RE = re.compile(r"^[a-z0-9][a-z0-9.\-]{1,38}[a-z0-9]$")
_SLUG_RE = re.compile(r"^[a-z0-9][a-z0-9\-]{1,38}[a-z0-9]$")
_TIME_RE = re.compile(r"^([01]?\d|2[0-3]):([0-5]\d)$")


@dataclass
class Issue:
    """One problem, located precisely enough for a school to fix it."""

    sheet: str
    row: int | None
    message: str
    fatal: bool = True

    def __str__(self) -> str:
        where = f"{self.sheet}" + (f" row {self.row}" if self.row else "")
        return f"{'ERROR' if self.fatal else 'warning'}  {where}: {self.message}"


@dataclass
class Plan:
    school: dict[str, Any] = field(default_factory=dict)
    periods: list[dict] = field(default_factory=list)
    breaks: list[dict] = field(default_factory=list)
    staff: list[dict] = field(default_factory=list)
    classes: list[dict] = field(default_factory=list)
    scopes: list[dict] = field(default_factory=list)
    slots: list[dict] = field(default_factory=list)
    students: list[dict] = field(default_factory=list)
    parents: list[dict] = field(default_factory=list)
    links: list[dict] = field(default_factory=list)
    issues: list[Issue] = field(default_factory=list)

    @property
    def errors(self) -> list[Issue]:
        return [i for i in self.issues if i.fatal]

    @property
    def warnings(self) -> list[Issue]:
        return [i for i in self.issues if not i.fatal]

    @property
    def ok(self) -> bool:
        return not self.errors

    def summary(self) -> str:
        return (
            f"{len(self.staff)} staff, {len(self.classes)} classes, "
            f"{len(self.scopes)} coordinator grants, {len(self.slots)} timetable slots, "
            f"{len(self.students)} students, {len(self.parents)} parents, "
            f"{len(self.links)} parent links"
        )


def _cells(ws) -> tuple[list[str], list[tuple[int, dict]]]:
    """(headings, [(excel_row_number, {heading: value})]) with example rows dropped."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headings = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: list[tuple[int, dict]] = []
    for n, raw in enumerate(rows[1:], start=2):
        vals = ["" if v is None else (v.strip() if isinstance(v, str) else v) for v in raw]
        if not any(str(v).strip() for v in vals):
            continue
        first = str(vals[0]).strip().upper()
        # The amber example row the school was told to delete. Tolerated rather
        # than rejected: a file is worth importing even if one row was missed.
        if first.startswith("EXAMPLE"):
            continue
        out.append((n, {h: v for h, v in zip(headings, vals) if h}))
    return headings, out


def _req(d: dict, heading: str) -> str:
    return str(d.get(heading, "") or "").strip()


def _email(v: str) -> str:
    return v.strip().lower()


def _yes(v: Any) -> bool | None:
    s = str(v or "").strip().lower()
    return True if s in {"yes", "y", "true", "1"} else False if s in {"no", "n", "false", "0"} else None


def _int(v: Any) -> int | None:
    try:
        return int(str(v).strip())
    except (TypeError, ValueError):
        return None


def parse_setup(path: str | Path, plan: Plan | None = None) -> Plan:
    """School, periods, breaks, staff, classes, coordinator scope, timetable."""
    from openpyxl import load_workbook

    plan = plan or Plan()
    wb = load_workbook(path, data_only=True)
    add = plan.issues.append

    def need(title: str):
        if title not in wb.sheetnames:
            add(Issue(title, None, "tab is missing — do not rename or delete tabs"))
            return None
        return wb[title]

    # ── 1. School (a vertical Field/Value form, not a grid) ──────────────────
    ws = need("1. School")
    if ws is not None:
        form = {
            str(r[0]).strip().rstrip("*").strip().lower(): r[1]
            for r in ws.iter_rows(min_col=1, max_col=2, values_only=True)
            if r[0] and not str(r[0]).strip().upper().startswith("EXAMPLE")
        }

        def f(label: str) -> Any:
            return form.get(label.lower())

        name = str(f("School name") or "").strip()
        slug = str(f("Preferred web address") or "").strip().lower()
        if not name:
            add(Issue("1. School", None, "School name is required"))
        if not slug:
            add(Issue("1. School", None, "Preferred web address is required"))
        elif not _SLUG_RE.match(slug):
            add(Issue("1. School", None,
                      f"Web address {slug!r} must be lowercase letters, numbers and hyphens only"))
        loc = str(f("Default interface language") or "en").strip() or "en"
        if loc not in LOCALES:
            add(Issue("1. School", None,
                      f"Default interface language {loc!r} is not one of: {', '.join(sorted(LOCALES))}"))
        for label in ("School day starts", "School day ends"):
            t = str(f(label) or "").strip()
            if t and not _TIME_RE.match(t):
                add(Issue("1. School", None, f"{label} {t!r} must be 24-hour HH:MM"))
        plan.school = {
            "name": name,
            "display_name": str(f("Display name") or "").strip() or None,
            "slug": slug,
            "country": str(f("Country") or "").strip().upper() or None,
            "locale": loc,
            "days": _int(f("Teaching days per week")) or 5,
            "start": str(f("School day starts") or "").strip() or None,
            "end": str(f("School day ends") or "").strip() or None,
            "period_minutes": _int(f("Length of one period, minutes")),
            "max_per_teacher_per_day": _int(f("Max periods per teacher per day")),
            "timetable_enabled": _yes(f("Enable timetable and cover plans")),
            "calendar": _yes(f("Enable calendar and notice board")),
            "school_analytics": _yes(f("Enable leadership analytics")),
            "school_assistant": _yes(f("Enable leadership briefing assistant")),
        }

    # ── 2/3. Periods and breaks ──────────────────────────────────────────────
    ws = need("2. Periods")
    if ws is not None:
        seen: set[int] = set()
        for n, r in _cells(ws)[1]:
            num = _int(r.get("Period number *"))
            label = _req(r, "Label *")
            at = _req(r, "Starts at *")
            if num is None:
                add(Issue("2. Periods", n, "Period number must be a whole number"))
                continue
            if num in seen:
                add(Issue("2. Periods", n, f"period {num} is listed twice"))
            seen.add(num)
            if at and not _TIME_RE.match(at):
                add(Issue("2. Periods", n, f"Starts at {at!r} must be 24-hour HH:MM"))
            plan.periods.append({"number": num, "label": label or f"P{num}", "time": at})
        plan.periods.sort(key=lambda p: p["number"])

    ws = need("3. Breaks")
    if ws is not None:
        for n, r in _cells(ws)[1]:
            after = _int(r.get("After period number *"))
            at = _req(r, "Starts at *")
            mins = _int(r.get("Minutes *"))
            if after is None or mins is None:
                add(Issue("3. Breaks", n, "After period number and Minutes must be whole numbers"))
                continue
            if at and not _TIME_RE.match(at):
                add(Issue("3. Breaks", n, f"Starts at {at!r} must be 24-hour HH:MM"))
            plan.breaks.append({"afterPeriod": after, "label": _req(r, "Label *") or "Break",
                                "time": at, "minutes": mins})

    # ── 4. Staff ─────────────────────────────────────────────────────────────
    ws = need("4. Staff")
    if ws is not None:
        for n, r in _cells(ws)[1]:
            email = _email(_req(r, "Email *"))
            name = _req(r, "Full name *")
            role = _req(r, "Role *").lower()
            if not name:
                add(Issue("4. Staff", n, "Full name is required"))
            if not email:
                add(Issue("4. Staff", n, "Email is required — it is how they sign in"))
            elif not _EMAIL_RE.match(email):
                add(Issue("4. Staff", n, f"{email!r} is not a valid email address"))
            elif email.endswith("@" + STUDENT_EMAIL_DOMAIN):
                add(Issue("4. Staff", n, f"{email!r} is a student sign-in address, not a staff email"))
            if role not in ROLES:
                add(Issue("4. Staff", n,
                          f"Role {role!r} must be one of: {', '.join(sorted(ROLES))}"))
            loc = _req(r, "Interface language")
            if loc and loc not in LOCALES:
                add(Issue("4. Staff", n, f"Interface language {loc!r} is not supported", fatal=False))
                loc = ""
            plan.staff.append({"row": n, "full_name": name, "email": email, "role": role,
                               "subjects": _req(r, "Subjects taught"),
                               "locale": loc or None, "staff_id": _req(r, "Staff ID")})
        _dupes(plan, "4. Staff", plan.staff, "email", "Email")

    # ── 5. Classes ───────────────────────────────────────────────────────────
    staff_emails = {s["email"] for s in plan.staff if s["email"]}
    ws = need("5. Classes")
    if ws is not None:
        for n, r in _cells(ws)[1]:
            cname = _req(r, "Class name *")
            grade = _req(r, "Grade *")
            temail = _email(_req(r, "Class teacher email *"))
            if not cname:
                add(Issue("5. Classes", n, "Class name is required"))
            if not grade:
                add(Issue("5. Classes", n, "Grade is required"))
            if not temail:
                add(Issue("5. Classes", n, "Class teacher email is required"))
            elif temail not in staff_emails:
                add(Issue("5. Classes", n,
                          f"class teacher {temail!r} is not on the Staff tab"))
            plan.classes.append({"row": n, "name": cname, "grade": grade,
                                 "teacher_email": temail, "room": _req(r, "Room")})
        _dupes(plan, "5. Classes", plan.classes, "name", "Class name")

    # ── 6. Coordinator scope ─────────────────────────────────────────────────
    coord_emails = {s["email"] for s in plan.staff if s["role"] == "coordinator"}
    ws = need("6. Coordinator Scope")
    if ws is not None:
        for n, r in _cells(ws)[1]:
            email = _email(_req(r, "Coordinator email *"))
            grade = _req(r, "Grade *")
            if not email or not grade:
                add(Issue("6. Coordinator Scope", n, "Coordinator email and Grade are required"))
                continue
            if email not in staff_emails:
                add(Issue("6. Coordinator Scope", n, f"{email!r} is not on the Staff tab"))
            elif email not in coord_emails:
                add(Issue("6. Coordinator Scope", n,
                          f"{email!r} is on the Staff tab but their Role is not 'coordinator'"))
            plan.scopes.append({"row": n, "email": email, "grade": grade,
                                "subject": _req(r, "Subject") or None})

    # ── 7. Timetable (optional) ──────────────────────────────────────────────
    class_names = {c["name"] for c in plan.classes if c["name"]}
    period_nums = {p["number"] for p in plan.periods}
    ws = wb["7. Timetable"] if "7. Timetable" in wb.sheetnames else None
    if ws is not None:
        occupied: dict[tuple[str, int, int], int] = {}
        teacher_day: dict[tuple[str, int], int] = {}
        for n, r in _cells(ws)[1]:
            cname = _req(r, "Class name *")
            day = _int(str(_req(r, "Day *")).split()[0]) if _req(r, "Day *") else None
            per = _int(r.get("Period number *"))
            subj = _req(r, "Subject *")
            temail = _email(_req(r, "Teacher email *"))
            if not cname or day is None or per is None or not subj or not temail:
                add(Issue("7. Timetable", n,
                          "Class name, Day, Period number, Subject and Teacher email are all required"))
                continue
            if cname not in class_names:
                add(Issue("7. Timetable", n, f"class {cname!r} is not on the Classes tab"))
            if temail not in staff_emails:
                add(Issue("7. Timetable", n, f"teacher {temail!r} is not on the Staff tab"))
            if not 1 <= day <= 7:
                add(Issue("7. Timetable", n, f"Day {day} must be 1-7"))
            if period_nums and per not in period_nums:
                add(Issue("7. Timetable", n, f"period {per} is not on the Periods tab"))
            key = (cname, day, per)
            if key in occupied:
                add(Issue("7. Timetable", n,
                          f"{cname} already has a lesson on day {day} period {per} "
                          f"(row {occupied[key]}) — a class cannot be in two places at once"))
            occupied[key] = n
            teacher_day[(temail, day)] = teacher_day.get((temail, day), 0) + 1
            plan.slots.append({"row": n, "class_name": cname, "day": day, "period": per,
                               "subject": subj, "teacher_email": temail, "room": _req(r, "Room")})
        cap = (plan.school or {}).get("max_per_teacher_per_day")
        if cap:
            for (email, day), count in sorted(teacher_day.items()):
                if count > cap:
                    add(Issue("7. Timetable", None,
                              f"{email} is timetabled for {count} periods on day {day}, "
                              f"above the school's stated maximum of {cap}", fatal=False))
    return plan


def parse_people(path: str | Path, plan: Plan | None = None) -> Plan:
    """Students, parents and the links between them. Optional file."""
    from openpyxl import load_workbook

    plan = plan or Plan()
    wb = load_workbook(path, data_only=True)
    add = plan.issues.append
    class_names = {c["name"] for c in plan.classes if c["name"]}

    ws = wb["1. Students"] if "1. Students" in wb.sheetnames else None
    if ws is None:
        add(Issue("1. Students", None, "tab is missing"))
    else:
        for n, r in _cells(ws)[1]:
            name = _req(r, "Full name *")
            user = _req(r, "Username *").lower()
            cname = _req(r, "Class name *")
            if not name:
                add(Issue("1. Students", n, "Full name is required"))
            if not user:
                add(Issue("1. Students", n, "Username is required — students sign in with it"))
            elif not _USERNAME_RE.match(user):
                add(Issue("1. Students", n,
                          f"username {user!r} must be lowercase letters, numbers, dots or hyphens, "
                          f"3-40 characters, no spaces"))
            if not cname:
                add(Issue("1. Students", n, "Class name is required"))
            elif class_names and cname not in class_names:
                add(Issue("1. Students", n,
                          f"class {cname!r} is not on the Classes tab of the School Setup file"))
            loc = _req(r, "Interface language")
            if loc and loc not in LOCALES:
                add(Issue("1. Students", n, f"Interface language {loc!r} is not supported", fatal=False))
                loc = ""
            plan.students.append({"row": n, "full_name": name, "username": user,
                                  "class_name": cname, "admission_no": _req(r, "Admission number"),
                                  "email": _email(_req(r, "Student email")) or None,
                                  "locale": loc or None})
        _dupes(plan, "1. Students", plan.students, "username", "Username")

    ws = wb["2. Parents"] if "2. Parents" in wb.sheetnames else None
    if ws is not None:
        for n, r in _cells(ws)[1]:
            email = _email(_req(r, "Email *"))
            name = _req(r, "Full name *")
            if not name:
                add(Issue("2. Parents", n, "Full name is required"))
            if not email:
                add(Issue("2. Parents", n, "Email is required — it is how the invitation reaches them"))
            elif not _EMAIL_RE.match(email):
                add(Issue("2. Parents", n, f"{email!r} is not a valid email address"))
            loc = _req(r, "Interface language")
            if loc and loc not in LOCALES:
                add(Issue("2. Parents", n, f"Interface language {loc!r} is not supported", fatal=False))
                loc = ""
            plan.parents.append({"row": n, "full_name": name, "email": email,
                                 "relationship": _req(r, "Relationship"),
                                 "phone": _req(r, "Phone"), "locale": loc or None})
        _dupes(plan, "2. Parents", plan.parents, "email", "Email")

    ws = wb["3. Parent Links"] if "3. Parent Links" in wb.sheetnames else None
    if ws is not None:
        pmails = {p["email"] for p in plan.parents if p["email"]}
        unames = {s["username"] for s in plan.students if s["username"]}
        pairs: set[tuple[str, str]] = set()
        for n, r in _cells(ws)[1]:
            email = _email(_req(r, "Parent email *"))
            user = _req(r, "Student username *").lower()
            if not email or not user:
                add(Issue("3. Parent Links", n, "Parent email and Student username are both required"))
                continue
            if email not in pmails:
                add(Issue("3. Parent Links", n, f"{email!r} is not on the Parents tab"))
            if user not in unames:
                add(Issue("3. Parent Links", n, f"{user!r} is not on the Students tab"))
            if (email, user) in pairs:
                add(Issue("3. Parent Links", n, "this parent and student are already linked above",
                          fatal=False))
                continue
            pairs.add((email, user))
            plan.links.append({"row": n, "parent_email": email, "username": user})
    return plan


def _dupes(plan: Plan, sheet: str, rows: list[dict], key: str, label: str) -> None:
    seen: dict[str, int] = {}
    for r in rows:
        v = str(r.get(key) or "").strip().lower()
        if not v:
            continue
        if v in seen:
            plan.issues.append(Issue(sheet, r.get("row"),
                                     f"{label} {v!r} is already used on row {seen[v]} — it must be unique"))
        else:
            seen[v] = r.get("row")


def report(plan: Plan) -> str:
    """What the founder reads before deciding to apply."""
    lines = [f"Parsed: {plan.summary()}"]
    if plan.school:
        s = plan.school
        lines.append(f"School: {s.get('name')!r} at school.sketchcast.app/{s.get('slug')}")
    for i in plan.errors:
        lines.append(str(i))
    for i in plan.warnings:
        lines.append(str(i))
    lines.append("")
    lines.append(
        f"{len(plan.errors)} error(s), {len(plan.warnings)} warning(s) — "
        + ("ready to apply" if plan.ok else "NOT importable until the errors are fixed")
    )
    return "\n".join(lines)
